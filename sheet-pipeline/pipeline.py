import os
import csv
import json
import re
from datetime import datetime, date

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SHEET_ID = os.environ.get("SHEET_ID", "")
OUT_DIR = os.environ.get("OUT_DIR", "/out")
START_ROW = int(os.environ.get("START_ROW", "5"))
SHEET_NAME = os.environ.get("SHEET_NAME")  # optional; default = active sheet

# Configure which columns contain linked text
LINK_COL_1 = os.environ.get("LINK_COL_1", "F").upper()
LINK_COL_2 = os.environ.get("LINK_COL_2", "G").upper()

# Configure output headers for those columns
LINK1_TEXT_HEADER = os.environ.get("LINK1_TEXT_HEADER", "timestamp 1")
LINK1_URL_HEADER  = os.environ.get("LINK1_URL_HEADER",  "timestamp 1 link")
LINK2_TEXT_HEADER = os.environ.get("LINK2_TEXT_HEADER", "ts 2")
LINK2_URL_HEADER  = os.environ.get("LINK2_URL_HEADER",  "ts 2 link")

if not SHEET_ID:
    raise SystemExit("Missing SHEET_ID env var")

os.makedirs(OUT_DIR, exist_ok=True)

xlsx_path = os.path.join(OUT_DIR, "sheet.xlsx")
csv_path  = os.path.join(OUT_DIR, "out.csv")
json_path = os.path.join(OUT_DIR, "out.json")

export_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

print("Downloading:", export_url)
r = requests.get(export_url, timeout=60)
r.raise_for_status()

# If not public, you may get HTML back.
ct = (r.headers.get("content-type") or "").lower()
if "text/html" in ct and r.content.lstrip().startswith(b"<!"):
    raise SystemExit("Got HTML (likely not public / needs auth).")

with open(xlsx_path, "wb") as f:
    f.write(r.content)

wb = load_workbook(xlsx_path, data_only=True)
ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

def cell_link(cell) -> str:
    # Whole-cell hyperlink target (works for typical linked text cells)
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target)
    return ""

def sanitize_header(h: str) -> str:
    h = (h or "").strip()
    if not h:
        return ""
    h = re.sub(r"\s+", " ", h)
    return h

def normalize_value(v):
    # Keep CSV nice and make JSON deterministic
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v

def json_default(o):
    if isinstance(o, (datetime, date)):
        return o.isoformat()
    # Fallback: avoid crashing on weird types
    return str(o)

# Determine max used column
max_col = ws.max_column

# Read header row (row 1) for all columns
base_headers = []
for c in range(1, max_col + 1):
    val = ws.cell(row=1, column=c).value
    hdr = sanitize_header(str(val)) if val is not None else ""
    if not hdr:
        hdr = get_column_letter(c)  # fallback to A/B/C...
    base_headers.append(hdr)

# Locate link columns
link_col1_idx = ws[f"{LINK_COL_1}1"].column
link_col2_idx = ws[f"{LINK_COL_2}1"].column

# Final headers: all non-link headers + (text,url) for each link column
final_headers = []
for i, hdr in enumerate(base_headers, start=1):
    if i == link_col1_idx or i == link_col2_idx:
        continue
    final_headers.append(hdr)

final_headers += [LINK1_TEXT_HEADER, LINK1_URL_HEADER, LINK2_TEXT_HEADER, LINK2_URL_HEADER]

rows = []
for r_idx in range(START_ROW, ws.max_row + 1):
    out = {}

    # Copy all non-link columns
    for c in range(1, max_col + 1):
        if c == link_col1_idx or c == link_col2_idx:
            continue
        hdr = base_headers[c - 1]
        val = ws.cell(row=r_idx, column=c).value
        out[hdr] = normalize_value(val)

    # Link columns (text + url)
    c1 = ws[f"{LINK_COL_1}{r_idx}"]
    c2 = ws[f"{LINK_COL_2}{r_idx}"]

    t1 = normalize_value(c1.value)
    u1 = cell_link(c1)
    t2 = normalize_value(c2.value)
    u2 = cell_link(c2)

    out[LINK1_TEXT_HEADER] = t1
    out[LINK1_URL_HEADER]  = u1
    out[LINK2_TEXT_HEADER] = t2
    out[LINK2_URL_HEADER]  = u2

    # Skip fully empty rows
    if any(v not in ("", None) for v in out.values()):
        rows.append(out)

with open(csv_path, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=final_headers, extrasaction="ignore")
    w.writeheader()
    w.writerows(rows)

with open(json_path, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2, default=json_default)

print("Wrote:", csv_path)
print("Wrote:", json_path)
